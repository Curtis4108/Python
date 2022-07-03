from optparse import Values
import tkinter as tk
from tkinter import Tk, ttk
import tkinter.messagebox
import customtkinter
from psutil import cpu_times_percent
import math
import pandas as pd
import numpy as np
import openpyxl

wb = openpyxl.load_workbook('CakeIngredients.xlsx')
sheet = wb.active

#Define cell locations
def get_ingredients(rownum, colnum, dozen:bool):
    ingredients = []
    measurements = []
    for row in range(rownum, 100):
        if(sheet.cell(row,colnum).value == 'Size'):
            size = sheet.cell(row, colnum+1).value
            if dozen == True:
                shape = "0"
                break
            shape = sheet.cell(row + 1, colnum + 1).value
            break
        ingredients.append(sheet.cell(row, colnum).value)
        measurements.append(sheet.cell(row, colnum + 1).value)
    Cake = {ingredients[i]:measurements[i] for i in range(len(ingredients))}
    return Cake, size, shape

def set_ingredients(rownum, colnum, ingredients, size, shape):
    for row in range(len(ingredients)):
        sheet.cell(rownum+row,colnum+1).value = ingredients[row]
    sheet.cell(rownum+row+1,colnum+1).value = size
    sheet.cell(rownum+row+2,colnum+1).value = shape
    wb.save('CakeIngredients.xlsx')
    return

storageLocations = {'ButterSpongeCake':{'row': 6, 'column': 2},
                    'VanillaSpongeCake':{'row': 6, 'column': 4},
                    'ChocolateSpongeCake':{'row': 6, 'column': 6},
                    'StrawberrySpongeCake':{'row': 6, 'column': 8},
                    'ChocolateMudCake':{'row': 17, 'column': 2},
                    'DarkChocolateMudCake':{'row': 17, 'column': 4},
                    'WhiteChocolateMudCake':{'row': 17, 'column': 6},
                    'FruitCake':{'row': 30, 'column': 2},
                    'CarrotCake':{'row': 30, 'column': 4},
                    'VanillaCupcake':{'row': 42, 'column': 2},
                    'ChocolateCupcake':{'row': 42, 'column': 4},
                    'StrawberryCupcake':{'row': 42, 'column': 6},
                    'SugarCookie':{'row': 51, 'column': 2},
                    'ChocolateChipCookie':{'row': 51, 'column': 4},
                    'ChocolateCookie':{'row': 51, 'column': 6},
                    'MnMCookie':{'row': 51, 'column': 8}}

ButterSpongeCake, BSC_size, BSC_shape = get_ingredients(storageLocations['ButterSpongeCake']['row'],storageLocations['ButterSpongeCake']['column'], False)
VanillaSpongeCake, VSC_size, VSC_shape = get_ingredients(storageLocations['VanillaSpongeCake']['row'],storageLocations['VanillaSpongeCake']['column'], False)
ChocolateSpongeCake, CSC_size, CSC_shape = get_ingredients(storageLocations['ChocolateSpongeCake']['row'],storageLocations['ChocolateSpongeCake']['column'], False)
StrawberrySpongeCake, SSC_size, SSC_shape = get_ingredients(storageLocations['StrawberrySpongeCake']['row'],storageLocations['StrawberrySpongeCake']['column'], False)
ChocolateMudCake, CMC_size, CMC_shape = get_ingredients(storageLocations['ChocolateMudCake']['row'],storageLocations['ChocolateMudCake']['column'], False)
DarkChocolateMudCake, DCMC_size, DCMC_shape = get_ingredients(storageLocations['DarkChocolateMudCake']['row'],storageLocations['DarkChocolateMudCake']['column'], False)
WhiteChocolateMudCake, WCMC_size, WCMC_shape = get_ingredients(storageLocations['WhiteChocolateMudCake']['row'],storageLocations['WhiteChocolateMudCake']['column'], False)
FruitCake, FC_size, FC_shape = get_ingredients(storageLocations['FruitCake']['row'],storageLocations['FruitCake']['column'], False)
CarrotCake, CC_size, CC_shape = get_ingredients(storageLocations['CarrotCake']['row'],storageLocations['CarrotCake']['column'], False)
VanillaCupcake, VCC_size, VCC_shape = get_ingredients(storageLocations['VanillaCupcake']['row'],storageLocations['VanillaCupcake']['column'], True)
ChocolateCupcake, CCC_size, CCC_shape = get_ingredients(storageLocations['ChocolateCupcake']['row'],storageLocations['ChocolateCupcake']['column'], True)
StrawberryCupcake, SCC_size, SCC_shape = get_ingredients(storageLocations['StrawberryCupcake']['row'],storageLocations['StrawberryCupcake']['column'], True)
SugarCookie, SCookie_size, SCookie_shape = get_ingredients(storageLocations['SugarCookie']['row'],storageLocations['SugarCookie']['column'], True)
ChocolateCookie, CCookie_size, CCookie_shape = get_ingredients(storageLocations['ChocolateCookie']['row'],storageLocations['ChocolateCookie']['column'], True)
ChocolateChipCookie, CCCookie_size, CCCookie_shape = get_ingredients(storageLocations['ChocolateChipCookie']['row'],storageLocations['ChocolateChipCookie']['column'], True)
MnMCookie, MnMCookie_size, MnMCookie_shape = get_ingredients(storageLocations['MnMCookie']['row'],storageLocations['MnMCookie']['column'], True)

massIngredients = ['Flour', 'Butter', 'Sugar', 'Cocoa Powder', 'Chocolate', 'Dark Chocolate', 'White Chocolate', 'Fruit', 'Nuts', 'M&Ms', 'Chocolate drops']
volumeIngredients = ['Water', 'Milk', 'Oil', 'Vanilla Extract', 'Strawberry Essence']

customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

allDropdowns = {'Cake':{'Sponge Cake': ['Vanilla', 'Butter', 'Chocolate', 'Strawberry'], 'Mud Cake': ['Dark Chocolate', 'Chocolate', 'White Chocolate'], 'Fruit Cake': ['Fruit Cake'], 'Carrot Cake': ['Carrot Cake']},
                'Cupcake':{"Vanilla", "Chocolate", "Strawberry"},
                'Cookie':{"Sugar", "Chocolate Chip", "Chocolate", "MnM"}}

cakeFlavours = {'Sponge Cake': ['Vanilla', 'Butter', 'Chocolate', 'Strawberry'], 'Mud Cake': ['Dark Chocolate', 'Chocolate', 'White Chocolate'], 'Fruit Cake': ['Fruit Cake'], 'Carrot Cake': ['Carrot Cake']}

cakeTypes = ['Sponge Cake', 'Mud Cake', 'Fruit Cake', 'Carrot Cake']

tin_size = [4, 6, 8, 9, 10, 12, 14]
round_surface = [12.56, 28.26, 50.24, 63.59, 78.5, 113.04, 153.86]
square_surface = [16, 36, 64, 81, 100, 144, 196]
round_cake_servings = [6, 14, 26, 30,38, 56, 76]
square_cake_servings = [8, 18, 32, 40, 50, 72, 98]
reverse_round_servings = sorted(round_cake_servings, reverse = True)
reverse_tin_sizes = sorted(tin_size, reverse = True)
reverse_square_servings = sorted(square_cake_servings, reverse = True)

class App(customtkinter.CTk):    
    WIDTH = 780
    HEIGHT = 520

    def __init__(self):
        super().__init__()
        self.title("Ingredients Estimator")
        self.geometry("600x600")
        self.tabsystem = ttk.Notebook(self)
        self.tab1 = ttk.Frame(self.tabsystem)
        self.tab2 = ttk.Frame(self.tabsystem)
        self.tab3 = ttk.Frame(self.tabsystem)
        self.tab4 = ttk.Frame(self.tabsystem)

        self.tabsystem.add(self.tab1, text='Cakes')
        self.tabsystem.add(self.tab2, text='Cupcakes')
        self.tabsystem.add(self.tab3, text='Cookies')
        self.tabsystem.add(self.tab4, text = 'Ingredients')

        #widgets for cake page
        self.labelCakeOptions = tk.Label(self.tab1, text="Choose cake type")
        self.optionmenu1Cakes = customtkinter.CTkOptionMenu(self.tab1, values=cakeTypes,command = self.cakeOptions)
        self.labelCakeFlavours = tk.Label(self.tab1, text="Choose cake flavour")
        self.optionmenu2Cakes = customtkinter.CTkOptionMenu(self.tab1, values=cakeFlavours["Sponge Cake"])
        self.labelCakeShape = tk.Label(self.tab1, text="Choose cake shape")
        self.optionmenu3Cakes = customtkinter.CTkOptionMenu(self.tab1, values=["Round", "Square"])
        self.labelCakeServings = tk.Label(self.tab1, text="Number of servings")
        self.entryCakeServings = customtkinter.CTkEntry(self.tab1, width=120,placeholder_text="Please enter a number")
        self.labelCakeCalculations = customtkinter.CTkLabel(self.tab1,text="Enter cake options and click the button to calculate measurements", height=200, width = 350,fg_color=("white", "gray38"), justify=tk.LEFT, wraplength = 320)
        self.labelCakeCalculations.text_label.place(relwidth = 1, relx = 0)
        self.CakeButton = customtkinter.CTkButton(self.tab1, text="Calculate", border_width=2, fg_color=None,  command=self.cake_calc)

        #add widgets to grid on cakes tab
        self.labelCakeOptions.grid(row = 1, column = 0)
        self.labelCakeCalculations.grid(row = 1, column = 2, rowspan = 12, padx = 10)
        self.optionmenu1Cakes.grid(row=3, column=0, pady=10, padx=10, sticky="w")
        self.labelCakeFlavours.grid(row = 5, column = 0)
        self.optionmenu2Cakes.grid(row=6, column=0, pady=10, padx=10, sticky="w")
        self.labelCakeShape.grid(row = 8, column = 0)
        self.optionmenu3Cakes.grid(row=9, column=0, pady=10, padx=10, sticky="w")
        self.labelCakeServings.grid(row = 11, column = 0)
        self.entryCakeServings.grid(row=12, column=0, columnspan=2, pady=10, padx=10, sticky="we")
        self.CakeButton.grid(row=17, column=0, columnspan=1, pady=10, padx=20, sticky="we")

        #widgets on Cupcake tab
        cupcakeFlavours = [0] * len(allDropdowns['Cupcake'])
        num = 0
        for i in allDropdowns['Cupcake']:
            cupcakeFlavours[num] = i
            num += 1
        self.labelCupcakeOptions = tk.Label(self.tab2, text="Choose cupcake flavour")
        self.optionmenu1Cupcakes = customtkinter.CTkOptionMenu(self.tab2, values=cupcakeFlavours, command=self.cakeOptions)
        self.labelCupcakeServings = tk.Label(self.tab2, text="Number of cupcakes")
        self.entryCupcakeServings = customtkinter.CTkEntry(self.tab2, width=120,placeholder_text="Please enter a number")
        self.labelCupcakeCalculations = customtkinter.CTkLabel(self.tab2,text="Enter cupcake options and click the button to calculate measurements", height=200, width = 350,fg_color=("white", "gray38"), justify=tk.LEFT, wraplength = 340)
        self.labelCupcakeCalculations.text_label.place(relwidth = 1, relx = 0)
        self.CupcakeButton = customtkinter.CTkButton(self.tab2, text="Calculate", border_width=2, fg_color=None,  command=self.scale_dozens_cupcakes)

        #add widgets to grid on cupcakes tab
        self.labelCupcakeOptions.grid(row = 1, column = 0)
        self.labelCupcakeCalculations.grid(row = 1, column = 2, rowspan = 12, padx = 10, pady=40)
        self.optionmenu1Cupcakes.grid(row=3, column=0, pady=10, padx=10, sticky="w")
        self.labelCupcakeServings.grid(row = 11, column = 0)
        self.entryCupcakeServings.grid(row=12, column=0, columnspan=2, pady=10, padx=10, sticky="we")
        self.CupcakeButton.grid(row=17, column=0, columnspan=1, pady=10, padx=20, sticky="we")

        #widgets on cookies tab
        cookieFlavours = [0] * len(allDropdowns['Cookie'])
        num = 0
        for i in allDropdowns['Cookie']:
            cookieFlavours[num] = i
            num += 1
        self.labelCookiesFlavours = tk.Label(self.tab3, text="Choose cookie flavour")
        self.optionmenuCookies = customtkinter.CTkOptionMenu(self.tab3, values = cookieFlavours)
        self.labelCookiesServings = tk.Label(self.tab3, text="Number of cookies")
        self.entryCookiesServings = customtkinter.CTkEntry(self.tab3, width=120,placeholder_text="Please enter a number")
        self.labelCookiesCalculations = customtkinter.CTkLabel(self.tab3,text="Enter cookie options and click the button to calculate measurements", height=200, width = 350,fg_color=("white", "gray38"), justify=tk.LEFT, wraplength = 350)
        self.labelCookiesCalculations.text_label.place(relwidth = 1, relx = 0)
        self.CookieButton = customtkinter.CTkButton(self.tab3, text="Calculate", border_width=2, fg_color=None,  command=self.scale_dozens_cookies)

        #add widgets to grid on cookies tab
        self.labelCookiesCalculations.grid(row = 2, column = 2, rowspan = 12, padx = 10, pady = 40)
        self.labelCookiesFlavours.grid(row = 3, column = 0)
        self.optionmenuCookies.grid(row=4, column=0, pady=10, padx=10, sticky="w")
        self.labelCookiesServings.grid(row = 6, column = 0)
        self.entryCookiesServings.grid(row=7, column=0, columnspan=2, pady=10, padx=10, sticky="we")
        self.CookieButton.grid(row=17, column=0, columnspan=1, pady=10, padx=20, sticky="we")

        self.createWidgets()

        self.tabsystem.pack(expand = 1, fill="both")

        # self.button = tk.Button(self, text='Get', command = self.on_button)
        # self.button_3 = tk.Button(self, text='Get', command = self.on_button)
        # self.label = customtkinter.CTkLabel(self,text="CTkLabel: Lorem ipsum dolor sit,\n", height=100, fg_color=("white", "gray38"), justify=tk.LEFT)
        # self.button.pack()
        # self.label.pack()
        # self.button_2.pack()
        # self.entry.pack()

    def cake_calc(self):
        servings = int(self.entryCakeServings.get())
        cake_shape = self.optionmenu3Cakes.get()
        used_tiers= []

        if cake_shape == "Round":
            tiers = 0
            if(servings <= round_cake_servings[0]):
                return
            for i in range(len(round_cake_servings)):
                if(servings >= reverse_round_servings[i]):
                    tiers += 1
                    servings -= reverse_round_servings[i]
                    used_tiers.append(reverse_tin_sizes[i])
        else:
            tiers = 0
            if(servings <= square_cake_servings[0]):
                return
            for i in range(len(square_cake_servings)):
                if(servings >= reverse_square_servings[i]):
                    tiers += 1
                    servings -= reverse_square_servings[i]
                    used_tiers.append(reverse_tin_sizes[i])
        self.scale_cake(used_tiers, cake_shape)
        return 

    def tiered_ingredients(self, ingredients, adjustment):
        new_list = ingredients
        for i in range(len(ingredients)):
            new_list[i] = float(ingredients [i]) * adjustment
        return new_list

    def scale_cake(self, tiers, shape):
        cakeType = self.optionmenu1Cakes.get()
        cakeFlavour = self.optionmenu2Cakes.get()
        CakeIngredients, cakeOriginalSize, cakeOriginalShape = self.defineDictionary('Cake', cakeFlavour, cakeType)
        adjustments = []
        if cakeOriginalShape == 'Round':
            original = round_surface[tin_size.index(int(cakeOriginalSize))]
        else:
            original = square_surface[tin_size.index(int(cakeOriginalSize))]
        for i in tiers:
            if shape =="Round":
                new_size = round_surface[tin_size.index(i)]
            else:
                new_size = square_surface[tin_size.index(i)]
            adjustments.append(new_size/original)

        original_ingredients = []
        ingredient_names = []
        for ingredients in CakeIngredients:
            original_ingredients.append(CakeIngredients[ingredients])
            ingredient_names.append(ingredients)

        ingredients_tiered = pd.DataFrame(original_ingredients, index = ingredient_names)

        for tierNum in range(len(tiers)):
            new_measurements = self.tiered_ingredients(original_ingredients, adjustments[tierNum])
            ingredients_tiered[tierNum+1] = new_measurements

        total = np.ceil(ingredients_tiered.iloc[:,1:].sum(axis=1))
        
        message = str(len(tiers)) + ' cakes of sizes '
        for i in tiers:
            message = message + str(i) + ' inches, '
        message = message[:-2] + '.\n'
        message = message + "Ingredients are: \n"
        for i in range(len(total)):
            message = message + ingredient_names[i] + ' - ' + str(total.iloc[i]) 
            if ingredient_names[i] in massIngredients:
                message = message + ' g'
            elif ingredient_names[i] in volumeIngredients:
                message = message + ' ml'
            message = message + '\n'
        self.labelCakeCalculations["text"] = message
        return

    def scale_dozens_cupcakes(self):
        Flavour = self.optionmenu1Cupcakes.get()
        number = int(self.entryCupcakeServings.get())
        Ingredients, OriginalSize, OriginalShape = self.defineDictionary('Cupcake', 0, Flavour)
        adjustments = number / 12

        original_ingredients = []
        ingredient_names = []
        for ingredients in Ingredients:
            original_ingredients.append(Ingredients[ingredients])
            ingredient_names.append(ingredients)

        new_measurements = self.tiered_ingredients(original_ingredients, adjustments)
        message = "Ingredients are: \n"
        for i in range(len(new_measurements)):
            message = message + ingredient_names[i] + ' - ' + str(np.ceil(new_measurements[i])) 
            if ingredient_names[i] in massIngredients:
                message = message + ' g'
            elif ingredient_names[i] in volumeIngredients:
                message = message + ' ml'
            message = message + '\n'
        self.labelCupcakeCalculations["text"] = message
        return

    def scale_dozens_cookies(self):
        Flavour = self.optionmenuCookies.get()
        number = int(self.entryCookiesServings.get())
        Ingredients, OriginalSize, OriginalShape = self.defineDictionary('Cookie', 0, Flavour)
        adjustments = number / 12

        original_ingredients = []
        ingredient_names = []
        for ingredients in Ingredients:
            original_ingredients.append(Ingredients[ingredients])
            ingredient_names.append(ingredients)

        new_measurements = self.tiered_ingredients(original_ingredients, adjustments)
        message = "Ingredients are: \n"
        for i in range(len(new_measurements)):
            message = message + ingredient_names[i] + ' - ' + str(np.ceil(new_measurements[i])) 
            if ingredient_names[i] in massIngredients:
                message = message + ' g'
            elif ingredient_names[i] in volumeIngredients:
                message = message + ' ml'
            message = message + '\n'
        self.labelCookiesCalculations["text"] = message
        return

    def getUpdateData(self, event):
        if self.DessertType.get() == 'Cake':
            self.TypeCombo['values'] = list(allDropdowns[self.DessertType.get()].keys())
            self.FlavourCombo.grid(row = 6, column = 0, pady = 10, padx = 10, sticky='w')
            self.CakeFlavour.grid(row = 5, column = 0)
            self.CakeFlavour["text"] = 'Select cake flavour'
            self.CakeType["text"] = 'Select cake type'
            self.TypeCombo.bind('<<ComboboxSelected>>', self.getUpdateData)
            self.FlavourCombo['values'] = list(allDropdowns[self.DessertType.get()][self.TypeCombo.get()])
        else:
            self.TypeCombo['values'] = list(allDropdowns[self.DessertType.get()])
            if self.DessertType.get() == 'Cupcake':
                self.CakeType["text"] = 'Select cupcake flavour'
                self.FlavourCombo.grid_forget()
                self.CakeFlavour.grid_forget()
            else:
                self.FlavourCombo.grid_forget()
                self.CakeFlavour.grid_forget()
        if self.TypeCombo.get() == "Fruit Cake":
            self.FlavourCombo.grid_forget()
            self.CakeFlavour.grid_forget()
        elif self.TypeCombo.get() == "Carrot Cake":
            self.FlavourCombo.grid_forget()
            self.CakeFlavour.grid_forget()
        else:
            self.FlavourCombo.set('')
        return

    def cakeOptions(self, choice):
        self.optionmenu2Cakes.configure(values= cakeFlavours[choice])
        if choice == "Fruit Cake" or choice == "Carrot Cake":
            self.optionmenu2Cakes.set(choice)
        else:
            self.optionmenu2Cakes.set('')
        return

    def createWidgets(self):
        self.DessertTypeLabel = tk.Label(self.tab4, text = 'Cake Type')
        self.DessertTypeLabel.grid(row = 0, column = 0)
        self.DessertType = ttk.Combobox(self.tab4, width = 15, values = list(allDropdowns.keys()))
        self.DessertType.bind('<<ComboboxSelected>>', self.getUpdateData)
        self.DessertType.grid(row=1, column=0, pady=10, padx=10, sticky="w")
        self.CakeType = tk.Label(self.tab4, text = 'Cake Type')
        self.CakeType.grid(row = 2, column = 0)
        self.CakeFlavour = tk.Label(self.tab4, text = 'Flavour')
        
        self.FlavourCombo = ttk.Combobox(self.tab4, width = 15)
        self.TypeCombo = ttk.Combobox(self.tab4, width = 15)
        self.TypeCombo.grid(row=3, column=0, pady=10, padx=10, sticky="w")

        self.createIngredients = customtkinter.CTkButton(self.tab4,text="Check Ingredients", border_width=2, fg_color=None,  command=self.build_ingredients_list)
        self.createIngredients.grid(row = 16, column = 0)

        self.updateIngredients = customtkinter.CTkButton(self.tab4,text="Update Ingredients", border_width=2, fg_color=None,  command=self.update_ingredients_list)
        self.updateIngredients.grid(row = 17, column = 0)

        return

    def defineDictionary(self, type, flavour, combo):
        if type == 'Cake':
            Description = flavour + combo
        else:
            Description = combo + type

        Description = Description.replace(" ", "")

        if Description == 'FruitCakeFruitCake':
            Description = 'FruitCake'
        elif Description == 'CarrotCakeCarrotCake':
            Description = 'CarrotCake'

        abbreviation = ''.join(c for c in Description if c.isupper())
        if type == 'Cupcake':
            abbreviation = abbreviation + 'C'
        elif type == 'Cookie':
            abbreviation = abbreviation + 'ookie'
            if abbreviation == 'MMCookie':
                abbreviation = 'MnMCookie'
        return globals()[Description], globals()[abbreviation + '_size'], globals()[abbreviation + '_shape']

    def storing_loc(self, type, flavour, combo):
        if type == 'Cake':
            Description = flavour + combo
        else:
            Description = combo + type
        Description = Description.replace(" ", "")
        return Description

    def build_ingredients_list(self):
        self.allIngredients = []
        FoodType = self.DessertType.get()
        FoodFlavour = self.FlavourCombo.get()
        FoodCombo = self.TypeCombo.get()
        CakeIngredients, cakeOriginalSize, cakeOriginalShape = self.defineDictionary(FoodType, FoodFlavour, FoodCombo)
        rownumber = 1
        for label in ttk.Frame.winfo_children(self.tab4):
            if label.grid_info():
                if int(label.grid_info()["column"])>1:
                    label.destroy()
        for ingredient in CakeIngredients.keys():            
            label = tk.Label(self.tab4,text = str(ingredient))
            label.grid(row=rownumber,column = 2, pady = 5)
            entry = tk.Entry(self.tab4, width = 10)
            entry.insert(tk.END, CakeIngredients[ingredient])
            entry.grid(row=rownumber,column = 3, pady = 5)
            if ingredient in massIngredients:
                labelText = 'grams'
            elif ingredient in volumeIngredients:
                labelText = 'ml'
            else:
                labelText = ' '
            label2 = tk.Label(self.tab4, text = labelText)
            label2.grid(row = rownumber, column = 4, pady = 5)
            rownumber += 1
            self.allIngredients.append(entry)
        if FoodType == 'Cake':
            labelCakeSize = tk.Label(self.tab4,text = "Tin Size")
            labelCakeSize.grid(row = rownumber+1,column = 2, pady = 5)
            self.entryCakeSize = tk.Entry(self.tab4, width = 10)
            self.entryCakeSize.insert(tk.END,str(cakeOriginalSize))
            self.entryCakeSize.grid(row=rownumber+1,column = 3, pady = 5)
            self.comboCakeShape = ttk.Combobox(self.tab4, values=['Round', 'Square'], width = 10)
            self.comboCakeShape.grid(row=rownumber+1,column = 4, pady = 5, sticky = 'w')
            if (cakeOriginalShape == 'Round'):
                self.comboCakeShape.current(0)
            else:
                self.comboCakeShape.current(1)
        return

    def update_ingredients_list(self):
        FoodType = self.DessertType.get()
        FoodFlavour = self.FlavourCombo.get()
        FoodCombo = self.TypeCombo.get()
        CakeIngredients, cakeOriginalSize, cakeOriginalShape = self.defineDictionary(FoodType, FoodFlavour, FoodCombo)
        newIngredients = []
        for i in range(len(self.allIngredients)):
            newIngredients.append(self.allIngredients[i].get())
        ingredientsList = list(CakeIngredients)
        for i in range(len(ingredientsList)):
            CakeIngredients[ingredientsList[i]] = newIngredients[i]
        cakeOriginalSize = self.entryCakeSize.get()
        cakeOriginalShape = self.comboCakeShape.get()

        description = self.storing_loc(FoodType, FoodFlavour, FoodCombo)
        rownum = storageLocations[description]['row']
        colnum = storageLocations[description]['column']
        set_ingredients(rownum, colnum, newIngredients, cakeOriginalSize, cakeOriginalShape)
        return

    def change_appearance_mode(self, new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def on_closing(self, event=0):
        self.destroy()


if __name__ == "__main__":
    app = App()
    app.mainloop()